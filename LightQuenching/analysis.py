# coding: utf-8
from ROOT import TFile, TH1F, TCanvas, TNtuple
import os
from openpyxl import Workbook
import csv
import re
#import matplotlib.pyplot as plt
import numpy as np

def GetRootFiles(path=os.getcwd(),gammaKey='Co60',neutronKey='neutron'):
  """ GetRootFiles

    Gets a list of root files, splitting up into gamma and neutrons
    according to gammaKey and neutronKey
  """
  gammaFiles = list()
  neutronFiles = list()
  for filename in os.listdir(path):
    [filename,ext] = os.path.splitext(filename)
    if ext == '.root':
      particleName = filename.split('_')[0].split('run')[0]
      filename = os.path.join(path,filename)+ext
      if particleName == gammaKey:
        gammaFiles.append(filename)
      elif particleName == neutronKey:
        neutronFiles.append(filename)
      else:
        print 'ROOT file '+filename+' is not reconized'
  return [gammaFiles,neutronFiles]


def GetThickness(filename):
  """ 
  GetThickness
  
  Parses a filename for a given thickness, returns the thickness in mm
  """
  [filename,ext] = os.path.splitext(filename)
  value = filename.split('_')[-1]
  # Uses a regular expression to seperate
  pattern = re.compile('(\d?.?\d+)(\w+)')
  tokens = pattern.split(value)
  if tokens[2] == 'm':
    return float(tokens[1])*1000
  elif tokens[2] == 'cm':
    return float(tokens[1])*10
  elif tokens[2] == 'mm':
    return float(tokens[1])
  elif tokens[2] == 'um':
    return float(tokens[1])*0.001
  else:
    print tokens[2]," is not a reconized prefix"
    raise Exception()

def PrintFiles(filelist,filename="EnergyDeposition.csv",
    tParse=GetThickness,histKey='eDepHist'):
  """ PlotFiles
  Prints the files files contained in filelist
  Keywords:
  """
  ws = wb.create_sheet()
  ws.title = sheetTitle
  row = 0
  col = 0
  longName = ['Energy','Frequency','Error']
  units =['MeV','Particles per Event','Particles per Event']
  for fname in filelist:
    f = TFile(fname,'r')
    print '\tSaving data from file: ',str(fname)
    hist = f.Get(histKey)
    # Writing the header
    i = 0
    for l,u in zip(longName,units):
      ws.cell(row=0,column=col+i).value = l
      ws.cell(row=1,column=col+i).value = u
      i += 1
    # Writing the Comments
    ws.cell(row=2,column=col+1).value = tParse(fname)
    # writing the data
    row = 3
    entries = hist.GetEntries()
    hist.Scale(1.0/entries)
    for i in range(int(entries)):
      ws.cell(row=row,column=col).value = hist.GetBinCenter(i)
      ws.cell(row=row,column=col+1).value = hist.GetBinContent(i)
      ws.cell(row=row,column=col+2).value = hist.GetBinError(i)
      row += 1
    col += 3

def PrintEDepSummary(gFiles,nFiles,wb=Workbook(),tParse=GetThickness,
  histKey='eDepHist'):
  """ PrintEDepSummary
  Prints the energy deposition summary
  """
  ws = wb.create_sheet()
  ws.title = 'EDepSummary'
  # Extrating the average values
  row = 0
  col = 0
  # Writing the headers
  longname = ['thickness','Average Energy Deposition','Avg. EDep Error']
  units =['mm','MeV','MeV']
  for l,u in zip(longname,units):
    ws.cell(row=0,column=col).value = l
    ws.cell(row=1,column=col).value = u
    ws.cell(row=0,column=col+3).value = l
    ws.cell(row=1,column=col+3).value = u
    col += 1
  ws.cell(row=2,column=1).value = 'Co-60'
  ws.cell(row=2,column=4).value = 'Cf-252'
  # Writing the data
  row = 3
  for fname in gFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    ws.cell(row=row,column=0).value = GetThickness(fname)
    ws.cell(row=row,column=1).value = hist.GetMean()
    ws.cell(row=row,column=2).value = hist.GetMeanError()
    row += 1
  row = 3
  for fname in nFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    ws.cell(row=row,column=3).value = GetThickness(fname)
    ws.cell(row=row,column=4).value = hist.GetMean()
    ws.cell(row=row,column=5).value = hist.GetMeanError()
    row += 1

def GetPosEDep(files,tParse=GetThickness,key='posEDepTuple'):
  """
  Returns a dictionary contaiting the mapping between the thickness (in mm)
  and an z, energy d
  """
  d = dict()
  # Looping through the files
  for fname in files:
    f = TFile(fname,'r')
    t = f.Get(key)
    entries = t.GetEntriesFast()
    data = np.zeros((2,entries))
    for entry in xrange(entries):
      # Loading entries into memory, addng to numpy array
      nb = t.GetEntry(entry)
      data[0][entry] += t.z
      data[1][entry] += t.EnergyDep
    d[tParse(fname)] = d
    print 'File: ',fname,' min z ',np.amin(data[0,:]),' max z ',np.amax(data[0,:])
  # Returning the dictonary
  return d

def EDepPosDist(d,csvFilename='EnergyDepDist.csv'):
  longname = ['Z position','Energy Depostion']
  units = ['mm','MeV']

def PlotEDepSummary(gFiles,nFiles,figureName='EDepSummary.png',tParse=GetThickness,
  histKey='eDepHist'):
  """ PlotEDepSummary
  Plotss the energy deposition summary
  """
  # Extrating the average values
  gT = list()
  gDep = list()
  gDepError = list()
  nT = list()
  nDep = list()
  nDepError = list()
  for fname in gFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    gT.append(GetThickness(fname))
    gDep.append(hist.GetMean())
    gDepError.append(hist.GetMeanError())
  for fname in nFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    nT.append(GetThickness(fname))
    nDep.append(hist.GetMean())
    nDepError.append(hist.GetMeanError())
  # Plotting
  plt.errorbar(gT,gDep,yerr=gDepError,fmt='ro')
  plt.hold(True)
  plt.errorbar(nT,nDep,yerr=nDepError,fmt='go')
  plt.xlabel("Thickness (mm)")
  plt.ylabel("Average Energy Deposition (MeV)")
  #plt.legend(["Co-60","Cf-252"])
  plt.xscale("log")
  plt.yscale("log")
  plt.grid(True)
  plt.savefig(figureName)

def main():
  print "Getting Files"
  [g,n] = GetRootFiles()
  print "Gamma Files: ",str(g)
  print "Neutron Files: ",str(n)
  print "Starting Data Analysis"
  GetPosEDep(g)
  GetPosEDep(n)

  wb = Workbook()
  #PrintEDepSummary(g,n,wb)
  #PlotEDepSummary(g,n)
  #print "Print Data"
  #PrintFiles(g,wb,"GammaEnergyDepostion")
  #PrintFiles(n,wb,"NeutronEnergyDepostion")
  #wb.save('EnergyDepAnalysis.xlsx')

if __name__ == "__main__":
  main()
